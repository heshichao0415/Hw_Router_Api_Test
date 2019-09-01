"""
# @Time    : 2019/8/29 19:32
# @Author  : 谢超
# @File    : do_excel.py
# @Function: 完成excel的读和写
"""
import openpyxl, sys
from Common import contants
from Common.do_log import MyLog


class Case:
    """
    把拥有同一特征的归为一个类
    测试用例类，每个测试用例，实际上就是它的一个实例
    """

    def __init__(self):
        self.case_id = None
        self.case_name = None
        self.method = None
        self.url = None
        self.data = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None


class DoExcel:
    """
    第一种存储数据的方法：
    1.循环遍历excel，获取到的每一行值赋值给Case类的实例中，存储在名为cases的列表中
    2.把所有属性放到一个字典中， 实例.__dict__
    3.和之前把读取到的数据以键值对存在一个列表中的方式一样
    4.这种方式是ddt数据驱动中以类属性来接收数据
    """

    # 初始化文件名和表单名，表示：实例化DoExcel的时候必须传file_name和sheet_name
    def __init__(self, file_name, sheet_name):

        self.file_name = file_name
        self.sheet_name = sheet_name

        # excle文件不存在，或者文件路径不正确就抛出异常，提高程序健壮性
        try:
            self.workbook = openpyxl.load_workbook(self.file_name)
        except Exception as e:
            MyLog(__name__).my_log().error('出错啦！！！！！！excle文件不存在，或者文件路径不正确。报错：{}'.format(e))

        # 表单名错误时抛出异常，提高程序健壮性
        try:
            self.sheet = self.workbook[self.sheet_name]
        except Exception as e:
            MyLog(__name__).my_log().error('出错啦！！！！！！excel的表单名错误。报错：{}'.format(e))

    def get_cases(self):
        """
        1.从ecxel第二行开始读取，没一行的数据存储到一个Case类属性中
        2.把类属性实例化，然后把Case对象存储到一个列表中，方便数据驱动
        :return: 返回一个列表，列表中存储Case类的对象
        """
        cases = []
        for row in range(2, self.sheet.max_row + 1):
            case = Case()
            case.case_id = self.sheet.cell(row=row, column=1).value
            case.case_name = self.sheet.cell(row=row, column=2).value
            case.method = self.sheet.cell(row=row, column=3).value
            case.url = self.sheet.cell(row=row, column=4).value
            case.data = self.sheet.cell(row=row, column=5).value
            case.expected = self.sheet.cell(row=row, column=6).value
            case.sql = self.sheet.cell(row=row, column=9).value
            cases.append(case)

        # 关闭excel
        self.workbook.close()
        return cases

    # 回写数据到excel第7列和第8列，如有需要可以进行添加
    def write_case(self, row, actual, result):
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(self.file_name)
        self.workbook.close()

if __name__ == '__main__':
    a = DoExcel(contants.case_file, "GetBlackBlackList")
    b = a.get_cases()
    print(b)


